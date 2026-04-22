"""Переменное число data-строк на YWJ1PF: 0 и 5."""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from src.core import config
from src.core.pipeline import run_pipeline

DATAR = config.DATA_START_ROW


def _ywj1_header_cols(ws) -> list[tuple[object, int]]:
    return [(ws.cell(2, c).value, c) for c in range(1, ws.max_column + 1)]


def _run_with_ywj1_row_count(example_workbook_path: Path, n_rows: int) -> None:
    import os

    fd, name = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    path = Path(name)
    shutil.copy2(example_workbook_path, path)
    for p in path.parent.glob(f"{path.stem}.backup_*.xlsx"):
        p.unlink(missing_ok=True)

    wb = load_workbook(path, data_only=False)
    wj = wb["YWJ1PF"]
    cols = _ywj1_header_cols(wj)
    template: dict[object, object] = {}
    first_data = DATAR
    for cname, cidx in cols:
        if cname is None:
            continue
        template[cname] = wj.cell(first_data, cidx).value
    for r in range(wj.max_row, DATAR - 1, -1):
        if r >= DATAR:
            wj.delete_rows(r, 1)
    if n_rows > 0:
        for i in range(n_rows):
            r = DATAR + i
            for cname, cidx in cols:
                if cname is None:
                    continue
                v = template.get(cname)
                if cname and "OTP" in str(cname):
                    wj.cell(r, cidx, value=str(i + 1))
                else:
                    wj.cell(r, cidx, value=v)
    wb.save(path)
    wb.close()

    try:
        run_pipeline(path)
    finally:
        path.unlink(missing_ok=True)
        for p in path.parent.glob(f"{path.stem}.backup_*.xlsx"):
            p.unlink(missing_ok=True)


def test_ywj1_zero_data_rows_pipeline(example_workbook_path) -> None:
    _run_with_ywj1_row_count(example_workbook_path, 0)


def test_ywj1_five_data_rows_pipeline(example_workbook_path) -> None:
    _run_with_ywj1_row_count(example_workbook_path, 5)
