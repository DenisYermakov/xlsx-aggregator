"""Оркестрация: загрузка книги, валидация, JOIN, запись, безопасное сохранение."""

from __future__ import annotations

import os
import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from src.core import joiner, writer
from src.core.account_extractor import extract_all_accounts
from src.core.config import (
    ALWAYS_APPEND,
    CONDITIONAL_APPEND,
    S5_SHEET,
    SC_SHEET,
    TARGET_SHEET,
    ACCOUNT_SOURCE_SHEETS,
    MAKE_BACKUP,
)
from src.core.sheet_reader import read_sheet_as_dicts, sheet_headers_list

ProgressFn = Callable[[int, str], None]


def _is_blank(v: object) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == ""


def _notify(progress: ProgressFn | None, pct: int, msg: str) -> None:
    if progress:
        progress(pct, msg)
    logger.info(msg)


def _assert_file_writable(path: Path) -> None:
    """Проверка блокировки файла (например, открыт в Excel) — до бэкапа и изменений."""
    try:
        with open(path, "ab"):
            pass
    except PermissionError as e:
        raise RuntimeError(
            f"Файл занят другим процессом (вероятно открыт в Excel): {path}. "
            "Закройте файл и повторите."
        ) from e


def _read_yw2pf_first_data_row_triggers(
    xlsx_path: Path, trigger_fields: set[str]
) -> dict[str, object]:
    """Читает значения триггеров из первой data-строки YW2PF (формулы — кэш data_only)."""
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        if TARGET_SHEET not in wb.sheetnames:
            return {}
        ws = wb[TARGET_SHEET]
        headers = [c.value for c in ws[2]]
        if not headers:
            return {}
        hmap = {h: i for i, h in enumerate(headers) if h is not None}
        row3 = [ws.cell(row=3, column=i + 1).value for i in range(len(headers))]
        out: dict[str, object] = {}
        for f in trigger_fields:
            if f in hmap:
                idx = hmap[f]
                if idx < len(row3):
                    out[f] = row3[idx]
        return out
    finally:
        wb.close()


def _headers_for_block(wb: Workbook, sheet_name: str) -> list:
    if sheet_name not in wb.sheetnames:
        return []
    return list(sheet_headers_list(wb, sheet_name))


def _safe_overwrite_save(wb: Workbook, original_path: str) -> dict[str, str]:
    """Сохраняет ``wb`` поверх ``original_path`` с бэкапом и атомарной подменой.

    Returns:
        ``{'result': путь_к_файлу, 'backup': путь_к_бэкапу}``
    """
    orig = Path(original_path)
    _assert_file_writable(orig)

    if MAKE_BACKUP:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = orig.with_name(f"{orig.stem}.backup_{ts}{orig.suffix}")
        shutil.copy2(orig, backup)
    else:
        backup = orig

    with tempfile.NamedTemporaryFile(
        dir=orig.parent,
        prefix=f".{orig.stem}.",
        suffix=".tmp",
        delete=False,
    ) as tf:
        tmp_path = Path(tf.name)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, orig)
    except Exception:
        if tmp_path.exists():
            tmp_path.unlink()
        raise
    if MAKE_BACKUP:
        return {"result": str(orig), "backup": str(backup)}
    return {"result": str(orig), "backup": str(backup)}


@dataclass
class PipelineResult:
    """Результат успешного прогона пайплайна."""

    result_path: str
    backup_path: str
    account_count: int


def run_pipeline(
    xlsx_path: str | Path,
    progress: ProgressFn | None = None,
) -> PipelineResult:
    """Выполняет полный цикл обработки и перезаписи файла.

    ТЗ: см. [PLAN.md] разд. 6–7 — загрузка, извлечение счетов, JOIN, запись,
    бэкап, атомарная замена.

    Args:
        xlsx_path: Путь к ``.xlsx`` (будет перезаписан).
        progress: Необязательный callback ``(процент, сообщение)``.

    Returns:
        ``PipelineResult`` с путями к файлу и бэкапу.

    Raises:
        RuntimeError: Файл занят, нет обязательных листов и т.п.
        OSError: Ошибка чтения/записи.
    """
    path = Path(xlsx_path).resolve()
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Ожидается файл .xlsx")

    _notify(progress, 0, "Старт обработки")
    _assert_file_writable(path)

    trigger_field_names = {c["trigger_field"] for c in CONDITIONAL_APPEND}
    triggers = _read_yw2pf_first_data_row_triggers(path, trigger_field_names)

    wb = load_workbook(path, data_only=False)
    try:
        if TARGET_SHEET not in wb.sheetnames:
            raise RuntimeError("В книге отсутствует обязательный лист YW2PF.")
        if SC_SHEET not in wb.sheetnames:
            raise RuntimeError("В книге отсутствует обязательный лист SCPF.")

        have_s5 = S5_SHEET in wb.sheetnames
        if not have_s5:
            logger.warning("Лист S5PF отсутствует — колонки S5* будут пустыми (штатно по ТЗ).")

        _notify(progress, 10, "Книга загружена")

        sc_rows = read_sheet_as_dicts(wb, SC_SHEET) if SC_SHEET in wb.sheetnames else []
        s5_rows = read_sheet_as_dicts(wb, S5_SHEET) if have_s5 else []

        # Данные для извлечения счетов: только листы-источники
        sheet_data: dict[str, list[dict]] = {}
        for sn in ACCOUNT_SOURCE_SHEETS:
            if sn not in wb.sheetnames:
                raise RuntimeError(f"Отсутствует лист, необходимый для счетов: {sn}")
            sheet_data[sn] = read_sheet_as_dicts(wb, sn)

        accounts = extract_all_accounts(sheet_data)
        _notify(progress, 30, f"Извлечено уникальных счетов: {len(accounts)}")

        acc_df: pd.DataFrame | None
        if not accounts:
            acc_df = None
        else:
            acc_df = joiner.build_account_table(accounts, sc_rows, s5_rows)
        _notify(progress, 50, "Таблица счетов построена")

        def _trigger_value(field: str) -> object:
            v = triggers.get(field)
            if v is None and sheet_data.get(TARGET_SHEET) and sheet_data[TARGET_SHEET]:
                v = sheet_data[TARGET_SHEET][0].get(field)
            return v

        conditional_include = {
            item["sheet"]: not _is_blank(_trigger_value(item["trigger_field"]))
            for item in CONDITIONAL_APPEND
        }

        ordered_blocks: list[tuple[str, list, list[dict]]] = []

        for ap_sheet in ALWAYS_APPEND:
            if ap_sheet not in wb.sheetnames:
                logger.warning("Лист %s отсутствует — блок пропущен.", ap_sheet)
                continue
            headers = _headers_for_block(wb, ap_sheet)
            rows = read_sheet_as_dicts(wb, ap_sheet)
            ordered_blocks.append((ap_sheet, headers, rows))

        for item in CONDITIONAL_APPEND:
            sh = item["sheet"]
            if not conditional_include.get(sh, False):
                logger.info("Условие для листа %s ложно — блок пропущен.", sh)
                continue
            if sh not in wb.sheetnames:
                logger.warning("Условный лист %s отсутствует — блок пропущен.", sh)
                continue
            headers = _headers_for_block(wb, sh)
            rows = read_sheet_as_dicts(wb, sh)
            ordered_blocks.append((sh, headers, rows))

        _notify(progress, 70, "Запись на YW2PF")
        writer.write_to_yw2pf(wb, ordered_blocks, acc_df)
        _notify(progress, 90, "Сохранение…")

        meta = _safe_overwrite_save(wb, str(path))
        _notify(progress, 100, "Готово")
        return PipelineResult(
            result_path=meta["result"],
            backup_path=meta["backup"],
            account_count=len(accounts),
        )
    finally:
        wb.close()


def main_cli() -> None:
    """Точка входа для ``python -m src.core.pipeline <файл.xlsx>``."""
    import sys

    if len(sys.argv) < 2:
        print("Usage: python -m src.core.pipeline <file.xlsx>", file=sys.stderr)
        raise SystemExit(1)
    p = Path(sys.argv[1])
    r = run_pipeline(p, progress=lambda n, t: print(f"{n}%: {t}"))
    print("OK", r)


if __name__ == "__main__":
    main_cli()
