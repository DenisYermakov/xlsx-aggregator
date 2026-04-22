"""Интеграционный прогон пайплайна на копии примера."""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from src.core.config import (
    TARGET_SHEET,
    ACCOUNTS_BLOCK_NAME,
    BLOCK_MARKER_PREFIX,
    BLOCK_MARKER_SUFFIX,
)
from src.core.pipeline import run_pipeline


def test_pipeline_runs_on_sample_copy(example_workbook_path) -> None:
    import os

    fd, name = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    path = Path(name)
    shutil.copy2(example_workbook_path, path)
    backup_glob = list(path.parent.glob(f"{path.stem}.backup_*.xlsx"))
    for b in backup_glob:
        b.unlink(missing_ok=True)
    try:
        r = run_pipeline(path)
        assert Path(r.result_path).exists()
        assert Path(r.backup_path).exists()
        assert Path(r.backup_path).name.endswith(".xlsx")

        w = load_workbook(path, data_only=False)
        try:
            ws = w[TARGET_SHEET]
            marks = [
                c.value
                for c in ws["A"]
                if isinstance(c.value, str) and c.value.startswith(BLOCK_MARKER_PREFIX)
            ]
            need = f"{BLOCK_MARKER_PREFIX}{ACCOUNTS_BLOCK_NAME}{BLOCK_MARKER_SUFFIX}"
            assert need in marks
        finally:
            w.close()
    finally:
        path.unlink(missing_ok=True)
        for p in path.parent.glob(f"{path.stem}.backup_*.xlsx"):
            p.unlink(missing_ok=True)
