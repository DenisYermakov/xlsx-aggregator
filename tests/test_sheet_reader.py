"""Тесты чтения листов в list[dict]."""

from __future__ import annotations

from openpyxl import load_workbook

from src.core.sheet_reader import read_sheet_as_dicts, sheet_headers_list


def test_read_ywj1_has_rows(example_workbook_path) -> None:
    wb = load_workbook(example_workbook_path, data_only=True)
    try:
        rows = read_sheet_as_dicts(wb, "YWJ1PF")
        assert len(rows) >= 1
        assert "YWJ1ANR" in rows[0] or len(rows[0]) > 0
    finally:
        wb.close()


def test_headers_from_empty_sheet_sanity(example_workbook_path) -> None:
    wb = load_workbook(example_workbook_path)
    try:
        h = sheet_headers_list(wb, "YW2PF")
        assert h is not None
        assert len(h) > 0
    finally:
        wb.close()
