"""Идемпотентность записи: повтор не дублирует блоки."""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.core.writer import write_to_yw2pf


def test_second_run_same_row_count(example_workbook_path) -> None:
    src = example_workbook_path
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    import os

    os.close(fd)
    path = Path(tmp)
    shutil.copy2(src, path)
    acc = pd.DataFrame(
        [
            {
                "SCAB": "1",
                "SCAN": "2",
                "SCAS": "3",
                "SCACT": None,
                "SCSAC": None,
                "SCNANC": None,
                "SCCCY": None,
                "SCBAL": None,
                "SCSUM0": None,
                "SCSUMD": None,
                "SCSUMC": None,
                "SCRBA": None,
                "S5BAL": None,
                "S5AIMD": 1.0,
                "S5AM1D": 2.0,
            }
        ]
    )
    blocks = [
        (
            "YW3PF",
            ["H1"],
            [{"H1": "v"}],
        )
    ]
    try:
        wb1 = load_workbook(path, data_only=False)
        write_to_yw2pf(wb1, blocks, acc)
        wb1.save(path)
        wb1.close()

        from src.core.config import TARGET_SHEET

        w1 = load_workbook(path)
        n1 = w1[TARGET_SHEET].max_row
        w1.close()
        wb2 = load_workbook(path, data_only=False)
        write_to_yw2pf(wb2, blocks, acc)
        wb2.save(path)
        wb2.close()

        w2 = load_workbook(path)
        n2 = w2[TARGET_SHEET].max_row
        w2.close()
        assert n1 == n2
    finally:
        path.unlink(missing_ok=True)
