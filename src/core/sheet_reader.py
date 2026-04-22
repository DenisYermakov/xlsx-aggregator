"""Чтение листов Excel в список словарей по строке заголовков (строка 2)."""

from __future__ import annotations

from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.config import DATA_START_ROW, HEADER_ROW


def read_headers(ws: Worksheet) -> list[str | None]:
    """Возвращает список имён полей из строки заголовков (строка HEADER_ROW)."""
    return [cell.value for cell in ws[HEADER_ROW]]


def read_sheet_as_dicts(wb: Workbook, sheet_name: str) -> list[dict[str, Any]]:
    """Возвращает все data-строки листа в виде списка словарей.

    ТЗ: листы-источники могут содержать произвольное число записей.
    В частности, на YWJ1PF их может быть 1, 2 или больше. Функция
    не делает предположений о количестве — читает всё, что есть.

    Формат листа (общий для всех листов PF в файле):
        строка 1: служебная ссылка «Go to Set Sheet» — игнорируется;
        строка 2: имена полей (заголовки);
        строка 3..N: данные.

    Args:
        wb: openpyxl Workbook.
        sheet_name: Имя листа, напр. ``"YWJ1PF"``.

    Returns:
        Список словарей вида ``[{field_name: value, ...}, ...]``, по одному
        на каждую непустую data-строку. Пустая строка (все ячейки None
        или пустые) пропускается (не прерывает чтение остальных строк).

    Example:
        >>> read_sheet_as_dicts(wb, "YWJ1PF")
        [{"YWJ1ANR": "F0ICRG20S210", "YWJ1OTP": "1", ...},
         {"YWJ1ANR": "F0ICRG20S210", "YWJ1OTP": "2", ...}]
    """
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[HEADER_ROW]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        # zip обрезает к длине headers — лишние ячейки в строке отбрасываются
        row_list = list(row)
        h_len = len(headers)
        if len(row_list) < h_len:
            row_list = row_list + [None] * (h_len - len(row_list))
        else:
            row_list = row_list[:h_len]
        rows.append(dict(zip(headers, row_list, strict=True)))
    return rows


def sheet_headers_list(wb: Workbook, sheet_name: str) -> list[str | None]:
    """Имена колонок листа (строка 2) — для пустого листа или перед чтением данных."""
    return read_headers(wb[sheet_name])
