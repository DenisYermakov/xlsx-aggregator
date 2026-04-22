"""Дозапись блоков и таблицы счетов на YW2PF, маркеры [XA:…], автофильтр."""

from __future__ import annotations

import pandas as pd
from loguru import logger
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.config import (
    ACCOUNTS_BLOCK_NAME,
    BLOCK_GAP,
    BLOCK_MARKER_PREFIX,
    BLOCK_MARKER_SUFFIX,
    ACCOUNT_COMPUTED_HEADER,
    TARGET_SHEET,
)

bold = Font(bold=True)


def _strip_previous_run(ws: Worksheet) -> None:
    """Удаляет все строки от пустой строки перед первым маркером ``[XA:`` вниз.

    ТЗ: идемпотентность — повторный запуск убирает хвост прошлой обработки
    (маркеры в колонке A).
    """
    first_marker_row: int | None = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        val = cell.value
        if isinstance(val, str) and val.startswith(BLOCK_MARKER_PREFIX):
            first_marker_row = cell.row
            break
    if first_marker_row is None:
        return
    # с первой пустой строки-зазора (строка перед маркером) до конца листа
    ws.delete_rows(first_marker_row - 1, ws.max_row - first_marker_row + 2)
    ws.auto_filter.ref = None


def _find_last_nonempty_row(ws: Worksheet) -> int:
    for r in range(ws.max_row, 0, -1):
        if any(
            ws.cell(row=r, column=c).value not in (None, "")
            for c in range(1, (ws.max_column or 0) + 1)
        ):
            return r
    return 0


def _apply_workbook_calc_flags(wb: Workbook) -> None:
    """Включает пересчёт при открытии в Excel (по плану §6.7)."""
    try:
        calc = wb.calculation
        if calc is not None:
            calc.fullCalcOnLoad = True  # type: ignore[assignment]
    except (AttributeError, TypeError) as e:
        logger.debug("Не удалось выставить fullCalcOnLoad: {}", e)
    try:
        props = wb.properties
        if props is not None and hasattr(props, "calcMode"):
            props.calcMode = "auto"  # type: ignore[assignment]
    except (AttributeError, TypeError) as e:
        logger.debug("Не удалось выставить calcMode: {}", e)


def write_to_yw2pf(
    wb: Workbook,
    ordered_blocks: list[tuple[str, list, list[dict]]],
    account_df: pd.DataFrame | None,
) -> None:
    """Пишет блоки листов и опционально таблицу счетов на ``YW2PF``.

    ТЗ: не трогать существующие строки выше; маркер ``[XA:`` — граница
    идемпотентности; таблица счетов с формулой ``-(S5AIMD+S5AM1D)``;
    автофильтр только на таблицу ``ACCOUNTS``.

    Args:
        wb: Книга (``data_only=False``).
        ordered_blocks: Список ``(имя_блока, headers, rows)`` — ``headers`` из строки 2
            исходного листа, ``rows`` — список dict по данным.
        account_df: Таблица из ``build_account_table``; если ``None`` или пустая —
            блок ``ACCOUNTS`` не пишется.
    """
    ws = wb[TARGET_SHEET]
    _strip_previous_run(ws)

    cursor = _find_last_nonempty_row(ws) + 1 + BLOCK_GAP

    for block_name, headers, rows in ordered_blocks:
        ws.cell(
            row=cursor,
            column=1,
            value=f"{BLOCK_MARKER_PREFIX}{block_name}{BLOCK_MARKER_SUFFIX}",
        ).font = bold
        cursor += 1
        for ci, h in enumerate(headers, start=1):
            ws.cell(row=cursor, column=ci, value=h).font = bold
        cursor += 1
        for row_dict in rows:
            for ci, h in enumerate(headers, start=1):
                ws.cell(row=cursor, column=ci, value=row_dict.get(h))
            cursor += 1
        cursor += BLOCK_GAP

    if account_df is None or account_df.empty:
        logger.info("Счетов нет — блок ACCOUNTS на YW2PF не записывается.")
    else:
        _write_account_table(ws, account_df, cursor)
    _apply_workbook_calc_flags(wb)


def _write_account_table(ws: Worksheet, account_df: pd.DataFrame, start_row: int) -> int:
    """Пишет ``[XA:ACCOUNTS]``, заголовки, данные, формулу, автофильтр."""
    cursor = start_row
    ws.cell(
        row=cursor,
        column=1,
        value=f"{BLOCK_MARKER_PREFIX}{ACCOUNTS_BLOCK_NAME}{BLOCK_MARKER_SUFFIX}",
    ).font = bold
    cursor += 1

    acc_headers = list(account_df.columns) + [ACCOUNT_COMPUTED_HEADER]
    header_row = cursor
    for ci, h in enumerate(acc_headers, start=1):
        ws.cell(row=cursor, column=ci, value=h).font = bold
    cursor += 1

    try:
        s5aimd_i = list(account_df.columns).index("S5AIMD") + 1
        s5am1d_i = list(account_df.columns).index("S5AM1D") + 1
    except ValueError as e:
        raise ValueError("В таблице счетов ожидаются колонки S5AIMD и S5AM1D.") from e
    s5aimd_letter = get_column_letter(s5aimd_i)
    s5am1d_letter = get_column_letter(s5am1d_i)
    neg_col_idx = len(account_df.columns) + 1

    for _, rec in account_df.iterrows():
        for ci, col in enumerate(account_df.columns, start=1):
            val = rec[col]
            out: object = None
            if not pd.isna(val):
                out = val
            ws.cell(row=cursor, column=ci, value=out)
        ws.cell(
            row=cursor,
            column=neg_col_idx,
            value=f"=-({s5aimd_letter}{cursor}+{s5am1d_letter}{cursor})",
        )
        cursor += 1
    data_end = cursor - 1
    if data_end < header_row:
        return cursor

    start_letter = get_column_letter(1)
    end_letter = get_column_letter(neg_col_idx)
    ws.auto_filter.ref = f"{start_letter}{header_row}:{end_letter}{data_end}"
    return cursor
